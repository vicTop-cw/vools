"""性能对比测试：vools.xl vs openpyxl"""
import os
import time
import tempfile

# 测试数据量
ROWS = 10000
COLS = 10

print(f'测试参数：{ROWS} 行 x {COLS} 列')
print('=' * 60)

# ========== vools.xl 写入测试 ==========
print('\n[vools.xl] 写入测试')
from vools.xl import Book

tmp_xl = os.path.join(tempfile.gettempdir(), 'perf_test_xl.xlsx')
start = time.time()

with Book() as book:
    sheet = book.add_sheet('Data')
    # trial 版本需要从第1行开始（避开A1）
    for row in range(1, ROWS + 1):
        for col in range(COLS):
            if col == 0:
                sheet.write_str(row, col, f'Name_{row}')
            elif col == 1:
                sheet.write_num(row, col, row * 100 + col)
            else:
                sheet.write_num(row, col, row + col)
    book.save(tmp_xl)

xl_write_time = time.time() - start
print(f'写入耗时: {xl_write_time:.3f} 秒')
print(f'文件大小: {os.path.getsize(tmp_xl)} bytes')

# ========== vools.xl 读取测试 ==========
print('\n[vools.xl] 读取测试')
start = time.time()

with Book() as book:
    book.load(tmp_xl)
    sheet = book.get_sheet(0)
    data = []
    for row in range(sheet.first_row, sheet.last_row + 1):
        row_data = []
        for col in range(sheet.first_col, sheet.last_col + 1):
            cell_type = sheet.cell_type(row, col)
            if cell_type == 1:  # 数字
                row_data.append(sheet.read_num(row, col))
            elif cell_type == 2:  # 字符串
                row_data.append(sheet.read_str(row, col))
            else:
                row_data.append(None)
        data.append(row_data)

xl_read_time = time.time() - start
print(f'读取耗时: {xl_read_time:.3f} 秒')
print(f'读取行数: {len(data)}')

# ========== openpyxl 测试 ==========
print('\n[openpyxl] 写入测试')
try:
    from openpyxl import Workbook
    tmp_openpyxl = os.path.join(tempfile.gettempdir(), 'perf_test_openpyxl.xlsx')
    start = time.time()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    for row in range(1, ROWS + 1):
        for col in range(1, COLS + 1):
            if col == 1:
                ws.cell(row=row, column=col, value=f'Name_{row-1}')
            elif col == 2:
                ws.cell(row=row, column=col, value=(row-1) * 100 + (col-1))
            else:
                ws.cell(row=row, column=col, value=(row-1) + (col-1))
    wb.save(tmp_openpyxl)

    openpyxl_write_time = time.time() - start
    print(f'写入耗时: {openpyxl_write_time:.3f} 秒')
    print(f'文件大小: {os.path.getsize(tmp_openpyxl)} bytes')

    # ========== openpyxl 读取测试 ==========
    print('\n[openpyxl] 读取测试')
    start = time.time()

    from openpyxl import load_workbook
    wb = load_workbook(tmp_openpyxl)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ROWS + 1, max_col=COLS):
        row_data = [cell.value for cell in row]
        data.append(row_data)

    openpyxl_read_time = time.time() - start
    print(f'读取耗时: {openpyxl_read_time:.3f} 秒')
    print(f'读取行数: {len(data)}')

    # ========== 对比结果 ==========
    print('\n' + '=' * 60)
    print('性能对比结果')
    print('=' * 60)
    print(f'写入速度: vools.xl {xl_write_time:.3f}s vs openpyxl {openpyxl_write_time:.3f}s')
    if openpyxl_write_time > 0:
        ratio = openpyxl_write_time / xl_write_time
        print(f'         vools.xl 快 {ratio:.1f}x')
    
    print(f'读取速度: vools.xl {xl_read_time:.3f}s vs openpyxl {openpyxl_read_time:.3f}s')
    if openpyxl_read_time > 0:
        ratio = openpyxl_read_time / xl_read_time
        print(f'         vools.xl 快 {ratio:.1f}x')

    # 清理
    os.remove(tmp_openpyxl)

except ImportError:
    print('openpyxl 未安装，跳过对比测试')

# 清理
os.remove(tmp_xl)
print('\n测试完成')
